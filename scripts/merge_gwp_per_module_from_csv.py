"""
merge_gwp_per_module_from_csv.py

Ingest per-module GWP A1–A3 (kgCO2e/module) from solar_database.csv
into INDICATORS_NORMALIZED of EPD_Hub_V3_PV_starter_enriched.xlsx.

Matching is done by module name (case/whitespace-insensitive).
"""

from pathlib import Path
import pandas as pd
import numpy as np

# -------------------------------------------------------------------
# CONFIG: adjust these to match YOUR filenames and CSV column names
# -------------------------------------------------------------------

BASE = Path(__file__).resolve().parent

# Your main workbook
XLSX_PATH = BASE / "EPD_Hub_V3_PV_starter_enriched.xlsx"

# Your CSV with per-module GWP values
CSV_PATH = BASE / "solar_database.csv"

# Sheet we update
NORM_SHEET = "INDICATORS_NORMALIZED"

# Column names IN THE CSV file:
CSV_NAME_COL = "Model_Series"                     # e.g. 'name', 'module_name', 'product_name'
CSV_GWP_MOD_COL = "GWP_A1A3_kgCO2e_per_Module"   # e.g. 'GWP_A1A3_per_module', change to your header

# Column name IN THE NORMALIZED sheet to fill:
NORM_NAME_COL = "name"
NORM_GWP_MOD_COL = "GWP_total_A1A3_per_module_kgCO2e"

# -------------------------------------------------------------------

def norm_name(s):
    """Normalize names: lowercase, strip, collapse whitespace."""
    if pd.isna(s):
        return ""
    return " ".join(str(s).strip().lower().split())


def is_zero_or_nan(x) -> bool:
    """Treat NaN and numeric 0 as 'empty' (safe to overwrite)."""
    if x is None or (isinstance(x, float) and np.isnan(x)):
        return True
    s = str(x).strip()
    if s == "":
        return True
    try:
        v = float(s.replace(",", "."))
        return v == 0.0
    except Exception:
        # non-numeric, treat as empty so we can overwrite with numeric
        return True


def to_float(x):
    try:
        return float(str(x).replace(",", "."))
    except Exception:
        return None


def main():
    # Safety checks
    if not XLSX_PATH.exists():
        raise SystemExit(f"Excel workbook not found: {XLSX_PATH}")
    if not CSV_PATH.exists():
        raise SystemExit(f"CSV file not found: {CSV_PATH}")

    print(f"Using Excel: {XLSX_PATH.name}")
    print(f"Using CSV  : {CSV_PATH.name}")

    # Load CSV
    df_csv = pd.read_csv(CSV_PATH)

    # Check required CSV columns
    for col in [CSV_NAME_COL, CSV_GWP_MOD_COL]:
        if col not in df_csv.columns:
            raise SystemExit(
                f"CSV column '{col}' not found. Available columns: {list(df_csv.columns)}"
            )

    # Keep only rows with a per-module value
    df_csv = df_csv[~df_csv[CSV_GWP_MOD_COL].isna()].copy()
    print(f"Rows in CSV with non-null per-module GWP: {len(df_csv)}")

    # Normalize name key in CSV
    df_csv["name_key"] = df_csv[CSV_NAME_COL].map(norm_name)

    # Build mapping name_key -> gwp_mod
    csv_map = {}
    for _, row in df_csv.iterrows():
        key = row["name_key"]
        if not key:
            continue
        gwp_mod = to_float(row[CSV_GWP_MOD_COL])
        if gwp_mod is None:
            continue
        # If duplicates exist, last one wins
        csv_map[key] = gwp_mod

    print(f"Unique name keys in CSV map: {len(csv_map)}")

    # Load normalized sheet
    df_norm = pd.read_excel(XLSX_PATH, sheet_name=NORM_SHEET)

    if NORM_NAME_COL not in df_norm.columns:
        raise SystemExit(f"Normalized sheet has no '{NORM_NAME_COL}' column.")

    # Ensure target column exists
    if NORM_GWP_MOD_COL not in df_norm.columns:
        df_norm[NORM_GWP_MOD_COL] = np.nan

    # Build name_key in normalized
    df_norm["name_key"] = df_norm[NORM_NAME_COL].map(norm_name)

    updated = 0
    total = len(df_norm)

    for idx, row in df_norm.iterrows():
        key = row["name_key"]
        if not key:
            continue
        if key not in csv_map:
            continue

        new_val = csv_map[key]
        cur_val = df_norm.at[idx, NORM_GWP_MOD_COL]

        # Only overwrite if current is empty or zero
        if is_zero_or_nan(cur_val):
            df_norm.at[idx, NORM_GWP_MOD_COL] = new_val
            updated += 1

    print(f"Rows in INDICATORS_NORMALIZED: {total}")
    print(f"Rows updated with per-module GWP: {updated}")

    # Drop helper key
    df_norm = df_norm.drop(columns=["name_key"])

    # Write back to same workbook (replace normalized sheet only)
    from openpyxl import load_workbook

    wb = load_workbook(XLSX_PATH)
    if NORM_SHEET in wb.sheetnames:
        # remove existing sheet and recreate
        ws = wb[NORM_SHEET]
        wb.remove(ws)
    ws_new = wb.create_sheet(NORM_SHEET)

    # Write df_norm to this sheet
    for j, col_name in enumerate(df_norm.columns, start=1):
        ws_new.cell(row=1, column=j, value=col_name)
    for i, (_, row) in enumerate(df_norm.iterrows(), start=2):
        for j, value in enumerate(row, start=1):
            ws_new.cell(row=i, column=j, value=value)

    wb.save(XLSX_PATH)
    print(f"Updated '{NORM_SHEET}' in workbook {XLSX_PATH.name}")


if __name__ == "__main__":
    main()